
import streamlit as st
from openpyxl import load_workbook
import tempfile

@st.cache_resource(show_spinner=False)
def load_workbook_cached(path: str):
    return load_workbook(path, data_only=False)

def bytes_to_tempfile(uploaded_file):
    data = uploaded_file.getbuffer()
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tf.write(data)
    tf.flush()
    tf.close()
    return tf.name
